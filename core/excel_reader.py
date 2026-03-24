"""Excel 读写模块 - 读写本地 xlsx 文件"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

logger = logging.getLogger(__name__)


def get_sheet_names(filepath: str | Path) -> list[str]:
    """返回 Excel 文件中的所有 Sheet 名称。"""
    wb = load_workbook(filepath, data_only=True, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def get_headers(filepath: str | Path, sheet_name: str) -> list[tuple[str, int]]:
    """返回指定 Sheet 的第 1 行列头列表。

    Returns:
        [(column_letter, col_index_1based), ...]
        例如 [('A', 1), ('B', 2), ...]
    """
    wb = load_workbook(filepath, data_only=True, read_only=True)
    ws = wb[sheet_name]
    headers: list[tuple[str, int]] = []
    for col_idx in range(1, ws.max_column + 1):
        letter = _col_letter(col_idx)
        cell_val = ws.cell(row=1, column=col_idx).value
        label = f"{letter}"
        if cell_val is not None:
            label += f" — {str(cell_val)[:30]}"
        headers.append((label, col_idx))
    wb.close()
    return headers


def read_data(
    filepath: str | Path,
    sheet_name: str,
    match_col: int,
    fill_cols: list[int],
    header_row: int = 1,
    data_start_row: int = 2,
) -> list[dict[str, Any]]:
    """从 Sheet 读取待填数据。

    Args:
        filepath: xlsx 路径
        sheet_name: Sheet 名称
        match_col: 匹配列 (1-based)，用于在钉钉表格中定位行
        fill_cols: 填入列列表 (1-based)，待填入的值
        header_row: 表头行号
        data_start_row: 数据起始行号

    Returns:
        [{'match_value': ..., 'fill_values': [str, ...], 'row': excel_row_number}, ...]
    """
    wb = load_workbook(filepath, data_only=True, read_only=True)
    ws = wb[sheet_name]
    rows = []

    for r in range(data_start_row, ws.max_row + 1):
        match_val = ws.cell(row=r, column=match_col).value
        if match_val is None:
            continue
        match_str = str(match_val).strip()
        if not match_str:
            continue

        values: list[str] = []
        for c in fill_cols:
            val = ws.cell(row=r, column=c).value
            if isinstance(val, datetime):
                val = val.strftime("%Y-%m-%d")
            elif val is None:
                val = ""
            values.append(str(val))
        rows.append({
            "match_value": match_str,
            "fill_values": values,
            "row": r,
        })

    wb.close()
    logger.info("从 %s 读取 %d 行数据", sheet_name, len(rows))
    return rows


def build_id_mapping(filepath: str | Path, sheet_name: str | None = None, id_col: int = 4) -> dict[str, int]:
    """从下载的钉钉表格 Excel 建立 id → 行号映射。

    Args:
        filepath: 下载的 xlsx 路径
        sheet_name: Sheet 名称，None 表示第一个 Sheet
        id_col: ID 所在列 (1-based)

    Returns:
        {id_value: excel_row_number}
    """
    wb = load_workbook(filepath, data_only=True, read_only=True)
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    mapping: dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=id_col).value
        if val:
            mapping[str(val).strip()] = r
    wb.close()
    logger.info("建立映射: %d 个 ID", len(mapping))
    return mapping


def _col_letter(col: int) -> str:
    """将 1-based 列号转为字母，如 1→A, 26→Z, 27→AA。"""
    result = ""
    while col > 0:
        col, remainder = divmod(col - 1, 26)
        result = chr(65 + remainder) + result
    return result
